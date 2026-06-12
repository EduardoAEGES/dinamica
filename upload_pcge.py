import os
import openpyxl
import requests
from dotenv import load_dotenv

# Cargar variables de entorno desde .env
load_dotenv()

SUPABASE_URL = os.getenv("SUPABASE_URL")
SUPABASE_KEY = os.getenv("SUPABASE_KEY")

if not SUPABASE_URL or not SUPABASE_KEY:
    print("Error: SUPABASE_URL o SUPABASE_KEY no están configurados en el archivo .env")
    exit(1)

EXCEL_FILE = "PCGE 2020 modificado rev.xlsx"
SHEET_NAME = "CATALOGO "

def clean_code(val):
    if val is None:
        return ""
    if isinstance(val, (int, float)):
        # Si es un número entero representado como float (ej. 10.0), convertir a int primero
        if val == int(val):
            return str(int(val)).strip()
        return str(val).strip()
    return str(val).strip()

def clean_desc(val):
    if val is None:
        return ""
    return str(val).strip()

def main():
    print(f"Abriendo el archivo {EXCEL_FILE}...")
    try:
        wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
    except Exception as e:
        print(f"Error al abrir el archivo Excel: {e}")
        return

    if SHEET_NAME not in wb.sheetnames:
        print(f"Error: La hoja '{SHEET_NAME}' no existe en el libro de trabajo. Hojas disponibles: {wb.sheetnames}")
        return

    sheet = wb[SHEET_NAME]
    records = []

    # Leer filas del excel
    for idx, row in enumerate(sheet.iter_rows(values_only=True), 1):
        # Omitir la fila de cabecera
        if idx == 1:
            continue
        
        raw_code = row[0]
        raw_desc = row[1]
        
        codigo = clean_code(raw_code)
        descripcion = clean_desc(raw_desc)
        
        # Omitir filas vacías
        if not codigo or not descripcion:
            continue
            
        # Omitir si es la cabecera por si acaso
        if codigo.lower() in ("código", "codigo") or descripcion.lower() in ("descripción de la cuenta", "descripcion de la cuenta"):
            continue

        nivel = len(codigo)
        records.append({
            "codigo": codigo,
            "descripcion": descripcion,
            "nivel": nivel
        })

    print(f"Se encontraron {len(records)} registros en el Excel.")

    # Eliminar duplicados en el lote basándose en el código primario (código único)
    unique_records = {}
    for r in records:
        unique_records[r["codigo"]] = r
    records = list(unique_records.values())
    print(f"Total registros únicos a subir: {len(records)}.")

    if len(records) > 0:
        print("Muestra de los primeros 5 registros:")
        for r in records[:5]:
            print(f"  Código: {r['codigo']} | Descripción: {r['descripcion']} | Nivel: {r['nivel']}")

    # Verificar si la tabla existe en Supabase y subir datos
    headers = {
        "apikey": SUPABASE_KEY,
        "Authorization": f"Bearer {SUPABASE_KEY}",
        "Content-Type": "application/json",
        "Prefer": "resolution=merge-duplicates" # Permite hacer Upsert si hay claves primarias
    }

    url = f"{SUPABASE_URL.rstrip('/')}/rest/v1/pcge_catalogo"

    print("\nVerificando conexión con Supabase...")
    try:
        # Hacemos una petición GET rápida para verificar si la tabla existe
        res = requests.get(f"{url}?limit=1", headers=headers)
        if res.status_code == 404:
            print("Error: La tabla 'pcge_catalogo' no existe en Supabase.")
            print("Debes crear la tabla primero en Supabase con el siguiente script SQL:")
            print("""
            CREATE TABLE pcge_catalogo (
                codigo TEXT PRIMARY KEY,
                descripcion TEXT NOT NULL,
                nivel INTEGER NOT NULL,
                created_at TIMESTAMPTZ DEFAULT NOW()
            );
            """)
            return
        elif res.status_code != 200 and res.status_code != 204:
            print(f"Error al conectar con Supabase (Código {res.status_code}): {res.text}")
            return
        else:
            print("¡Conexión exitosa! La tabla 'pcge_catalogo' existe.")
    except Exception as e:
        print(f"Error de red al conectar con Supabase: {e}")
        return

    # Subir en lotes de 100 registros
    batch_size = 100
    print(f"\nSubiendo {len(records)} registros en lotes de {batch_size}...")
    
    for i in range(0, len(records), batch_size):
        batch = records[i:i + batch_size]
        try:
            # Enviamos el lote en formato JSON
            # Nota: Usamos POST con Prefer: resolution=merge-duplicates para hacer Upsert basado en la clave primaria 'codigo'
            # En PostgREST, para hacer Upsert, la clave primaria o restricción única debe especificarse o inferirse.
            # PostgREST requiere el encabezado 'Prefer: resolution=merge-duplicates' y que se envíen los campos de clave primaria.
            res = requests.post(url, json=batch, headers=headers)
            if res.status_code in (200, 201, 204):
                print(f"Lote {i//batch_size + 1} subido exitosamente (registros {i+1} a {min(i+batch_size, len(records))}).")
            else:
                print(f"Error al subir lote {i//batch_size + 1}: {res.status_code} - {res.text}")
        except Exception as e:
            print(f"Excepción al subir lote {i//batch_size + 1}: {e}")

    print("\nProceso finalizado.")

if __name__ == "__main__":
    main()
